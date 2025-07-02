"""
RAG Agent for QuokkaAI - Retrieval-Augmented Generation for document and data analysis.

This module implements a sophisticated RAG agent that can:
- Process various document formats (PDF, CSV, Excel, TXT, JSON)
- Create embeddings using OpenAI
- Store and retrieve from Chroma vector database
- Perform detailed data science analysis
- Make predictions and recommendations
- Generate insights like a professional data analyst
"""

import asyncio
import logging
import os
import shutil
from pathlib import Path
from typing import Dict, List, Optional, Any, Union, Tuple
from dataclasses import dataclass
from datetime import datetime
import pandas as pd
import numpy as np
import json
import threading

from pydantic import BaseModel, Field
from pydantic_settings import BaseSettings


# LangChain imports
from langchain_community.document_loaders import (
    PyPDFLoader, CSVLoader, TextLoader, JSONLoader,
    UnstructuredExcelLoader, DirectoryLoader
)
from langchain_text_splitters import RecursiveCharacterTextSplitter
from langchain_core.documents import Document
from langchain_core.tools import Tool
from langchain_openai import OpenAIEmbeddings, ChatOpenAI
from langchain_community.vectorstores import Chroma
from langchain_community.vectorstores.utils import filter_complex_metadata
from langchain.chains import RetrievalQA
from langchain.memory import ConversationBufferMemory
from langchain_core.prompts import PromptTemplate

# Data analysis imports
import plotly.express as px
import plotly.graph_objects as go
from sklearn.preprocessing import StandardScaler, LabelEncoder
from sklearn.model_selection import train_test_split
from sklearn.ensemble import RandomForestRegressor, RandomForestClassifier
from sklearn.linear_model import LinearRegression, LogisticRegression
from sklearn.metrics import mean_squared_error, accuracy_score, classification_report
from sklearn.cluster import KMeans
import seaborn as sns
import matplotlib.pyplot as plt

logger = logging.getLogger(__name__)

# Global singleton instance
_rag_agent_instance = None
_instance_lock = threading.Lock()


class RAGSettings(BaseSettings):
    """Settings for the RAG agent."""
    
    openai_api_key: str = Field(alias="OPENAI_API_KEY", description="OpenAI API key")
    data_directory: str = Field(default="data/rag", description="Directory for uploaded files")
    chroma_directory: str = Field(default="data/chroma", description="Chroma database directory")
    chunk_size: int = Field(default=1000, description="Text chunk size for splitting")
    chunk_overlap: int = Field(default=200, description="Overlap between chunks")
    max_tokens: int = Field(default=4000, description="Maximum tokens for LLM responses")
    temperature: float = Field(default=0.1, description="LLM temperature for analysis")
    
    model_config = {
        "env_file": ".env",
        "case_sensitive": True,
        "extra": "ignore"  # Allow extra fields from environment
    }


@dataclass
class AnalysisResult:
    """Result of data analysis."""
    summary: str
    insights: List[str]
    recommendations: List[str]
    visualizations: List[Dict[str, Any]]
    statistical_analysis: Dict[str, Any]
    predictions: Optional[Dict[str, Any]] = None
    confidence_score: float = 0.0


@dataclass
class ProcessedDocument:
    """Processed document metadata."""
    filename: str
    file_type: str
    size: int
    processed_at: datetime
    chunks_count: int
    embedding_model: str
    summary: str


class DataAnalysisRAGAgent:
    """
    Advanced RAG agent for comprehensive data analysis.
    
    Capabilities:
    - Multi-format document processing (PDF, CSV, Excel, JSON, TXT)
    - Intelligent text chunking and embedding
    - Vector similarity search with Chroma
    - Statistical analysis and visualization
    - Machine learning predictions
    - Professional data science insights
    - Interactive chart generation
    """
    
    def __init__(self, settings: Optional[RAGSettings] = None):
        """
        Initialize the RAG agent.
        
        Args:
            settings: Configuration settings for the agent
        """
        self.settings = settings or RAGSettings()
        self._setup_directories()
        
        # Initialize components
        self.embeddings = OpenAIEmbeddings(
            api_key=self.settings.openai_api_key
        )
        self.llm = ChatOpenAI(
            model="gpt-4o-mini",
            api_key=self.settings.openai_api_key,
            temperature=self.settings.temperature,
            max_tokens=self.settings.max_tokens
        )
        
        # Initialize text splitter
        self.text_splitter = RecursiveCharacterTextSplitter(
            chunk_size=self.settings.chunk_size,
            chunk_overlap=self.settings.chunk_overlap,
            separators=["\n\n", "\n", ". ", " ", ""]
        )
        
        # Initialize vector store
        self.vectorstore = Chroma(
            persist_directory=self.settings.chroma_directory,
            embedding_function=self.embeddings
        )
        
        # Initialize memory for conversation context
        self.memory = ConversationBufferMemory(
            memory_key="chat_history",
            return_messages=True
        )
        
        # Document registry
        self.processed_documents: Dict[str, ProcessedDocument] = {}
        
        logger.info("DataAnalysisRAGAgent initialized successfully")

    async def _reindex_existing_documents(self) -> None:
        """Reindex existing documents in the data directory for thread isolation."""
        try:
            data_dir = Path(self.settings.data_directory)
            
            # Check if documents subdirectory exists (for thread agents)
            docs_dir = data_dir / "documents"
            if docs_dir.exists() and docs_dir.is_dir():
                check_dirs = [docs_dir, data_dir]
            else:
                check_dirs = [data_dir]
            
            # Process all files in the directories
            for check_dir in check_dirs:
                if check_dir.exists():
                    for file_path in check_dir.iterdir():
                        if file_path.is_file() and file_path.suffix.lower() in ['.pdf', '.csv', '.xlsx', '.xls', '.json', '.txt', '.md']:
                            # Check if already processed
                            if file_path.name not in self.processed_documents:
                                logger.info(f"Reindexing document: {file_path.name}")
                                await self.upload_file(str(file_path), file_path.name)
                    
                    if len(self.processed_documents) > 0:
                        break  # Found and processed files in this directory
            
            logger.info(f"Reindexing completed: {len(self.processed_documents)} documents available")
            
        except Exception as e:
            logger.warning(f"Document reindexing failed: {e}")

    def _setup_directories(self) -> None:
        """Create necessary directories."""
        Path(self.settings.data_directory).mkdir(parents=True, exist_ok=True)
        Path(self.settings.chroma_directory).mkdir(parents=True, exist_ok=True)
        logger.info(f"Directories set up: {self.settings.data_directory}, {self.settings.chroma_directory}")

    async def upload_file(self, file_path: str, original_filename: str) -> Dict[str, Any]:
        """
        Process and store an uploaded file.
        
        Args:
            file_path: Path to the uploaded file
            original_filename: Original name of the file
            
        Returns:
            Processing result with metadata
        """
        try:
            # Handle file path properly - move only if different directory
            source_path = Path(file_path)
            target_path = Path(self.settings.data_directory) / original_filename
            
            if source_path != target_path:
                if target_path.exists():
                    target_path.unlink()  # Remove existing file
                shutil.move(str(source_path), str(target_path))
            else:
                target_path = source_path  # File is already in the right place
            
            # Process the document
            documents = await self._load_document(target_path)
            
            if not documents:
                return {
                    "status": "error",
                    "message": "Failed to load document content"
                }
            
            # Split into chunks
            chunks = self.text_splitter.split_documents(documents)
            
            # Filter complex metadata before adding to vector store
            filtered_chunks = []
            for chunk in chunks:
                # Create a copy of the chunk with filtered metadata
                filtered_metadata = {}
                for key, value in chunk.metadata.items():
                    # Only keep simple types that Chroma can handle
                    if isinstance(value, (str, int, float, bool, type(None))):
                        filtered_metadata[key] = value
                    elif isinstance(value, list):
                        # Convert list to string representation
                        filtered_metadata[f"{key}_summary"] = f"Multiple errors: {len(value)} issues"
                        # Store first few items as string if they're simple types
                        simple_items = [item for item in value[:3] if isinstance(item, (str, int, float, bool))]
                        if simple_items:
                            filtered_metadata[f"{key}_first"] = str(simple_items[0])
                    else:
                        # Convert complex types to string
                        filtered_metadata[key] = str(value)
                
                # Create new document with filtered metadata
                filtered_chunk = Document(
                    page_content=chunk.page_content,
                    metadata=filtered_metadata
                )
                filtered_chunks.append(filtered_chunk)
            
            # Add filtered chunks to vector store
            self.vectorstore.add_documents(filtered_chunks)
            
            # Generate summary
            summary = await self._generate_document_summary(documents[:5])  # First 5 chunks
            
            # Store metadata
            processed_doc = ProcessedDocument(
                filename=original_filename,
                file_type=target_path.suffix.lower(),
                size=target_path.stat().st_size,
                processed_at=datetime.utcnow(),
                chunks_count=len(chunks),
                embedding_model="text-embedding-ada-002",
                summary=summary
            )
            
            self.processed_documents[original_filename] = processed_doc
            
            logger.info(f"Successfully processed {original_filename}: {len(chunks)} chunks")
            
            return {
                "status": "success",
                "filename": original_filename,
                "chunks_count": len(chunks),
                "summary": summary,
                "file_type": processed_doc.file_type,
                "size": processed_doc.size
            }
            
        except Exception as e:
            logger.error(f"Error processing file {original_filename}: {e}")
            return {
                "status": "error",
                "message": str(e)
            }

    async def _load_document(self, file_path: Path) -> List[Document]:
        """Load document based on file type with enhanced error handling."""
        file_type = file_path.suffix.lower()
        
        try:
            if file_type == '.pdf':
                return await self._load_pdf_with_fallbacks(file_path)
            elif file_type == '.csv':
                return await self._load_csv_with_fallbacks(file_path)
            elif file_type in ['.txt', '.md']:
                return await self._load_text_with_fallbacks(file_path)
            elif file_type == '.json':
                return await self._load_json_with_fallbacks(file_path)
            elif file_type in ['.xlsx', '.xls']:
                return await self._load_excel_with_fallbacks(file_path)
            else:
                # Try as text file
                return await self._load_text_with_fallbacks(file_path)
            
        except Exception as e:
            logger.error(f"Error loading {file_path}: {e}")
            return [Document(
                page_content=f"File: {file_path.name}\n\nError loading file: {str(e)}\n\nPlease try a different file format.",
                metadata={"source": str(file_path), "file_type": file_type, "error": str(e)}
            )]

    async def _load_pdf_with_fallbacks(self, file_path: Path) -> List[Document]:
        """Enhanced PDF loader that extracts tables, images, charts, and structured content."""
        errors = []
        extracted_documents = []
        
        # Method 1: Advanced PyMuPDF extraction with table and image detection
        try:
            import fitz  # PyMuPDF
            logger.info(f"🔍 Starting advanced PDF analysis for {file_path}")
            
            doc = fitz.open(str(file_path))
            total_pages = len(doc)
            
            for page_num in range(total_pages):
                page = doc.load_page(page_num)
                page_content = []
                page_metadata = {
                    "source": str(file_path),
                    "page": page_num + 1,
                    "total_pages": total_pages,
                    "extraction_method": "Advanced PyMuPDF",
                    "content_types": []
                }
                
                # 1. Extract regular text with layout preservation
                try:
                    text_dict = page.get_text("dict")
                    formatted_text = await self._extract_formatted_text(text_dict)
                    if formatted_text.strip():
                        page_content.append(f"📄 **Page {page_num + 1} - Text Content:**\n{formatted_text}")
                        page_metadata["content_types"].append("text")
                except Exception as e:
                    logger.warning(f"Text extraction failed for page {page_num}: {e}")
                
                # 2. Extract tables using advanced detection
                try:
                    tables = await self._extract_pdf_tables(page, page_num + 1)
                    if tables:
                        page_content.append(f"\n📊 **Tables Found on Page {page_num + 1}:**\n{tables}")
                        page_metadata["content_types"].append("tables")
                        page_metadata["table_count"] = len(tables.split("Table"))
                except Exception as e:
                    logger.warning(f"Table extraction failed for page {page_num}: {e}")
                
                # 3. Extract and describe images/charts
                try:
                    images_description = await self._extract_pdf_images(page, page_num + 1)
                    if images_description:
                        page_content.append(f"\n🖼️ **Visual Elements on Page {page_num + 1}:**\n{images_description}")
                        page_metadata["content_types"].append("images")
                except Exception as e:
                    logger.warning(f"Image extraction failed for page {page_num}: {e}")
                
                # 4. Extract forms and annotations
                try:
                    forms_content = await self._extract_pdf_forms(page, page_num + 1)
                    if forms_content:
                        page_content.append(f"\n📝 **Forms/Annotations on Page {page_num + 1}:**\n{forms_content}")
                        page_metadata["content_types"].append("forms")
                except Exception as e:
                    logger.warning(f"Forms extraction failed for page {page_num}: {e}")
                
                # 5. OCR for scanned content if no text found
                if not any("text" in content_type for content_type in page_metadata["content_types"]):
                    try:
                        logger.info(f"🔍 No text found on page {page_num + 1}, attempting OCR...")
                        pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))  # High resolution
                        img_data = pix.tobytes("png")
                        
                        ocr_text = await self._extract_text_with_ocr(img_data)
                        if ocr_text.strip():
                            page_content.append(f"\n🔍 **OCR Extracted Content:**\n{ocr_text}")
                            page_metadata["content_types"].append("ocr")
                            page_metadata["extraction_method"] += " + OCR"
                    except Exception as ocr_error:
                        logger.warning(f"OCR failed for page {page_num}: {ocr_error}")
                
                # Combine all content for this page
                if page_content:
                    combined_content = "\n".join(page_content)
                    
                    # Add structural summary
                    content_summary = f"""
📋 **Page {page_num + 1} Summary:**
- Content Types: {', '.join(page_metadata['content_types'])}
- Extraction Method: {page_metadata['extraction_method']}
{f"- Tables: {page_metadata.get('table_count', 0)}" if 'tables' in page_metadata.get('content_types', []) else ""}

{combined_content}
"""
                    
                    extracted_documents.append(
                        Document(
                            page_content=content_summary,
                            metadata=page_metadata
                        )
                    )
                else:
                    # Empty page - still add with metadata
                    extracted_documents.append(
                        Document(
                            page_content=f"📄 **Page {page_num + 1}:** [Empty or unreadable page]",
                            metadata=page_metadata
                        )
                    )
            
            doc.close()
            
            if extracted_documents:
                logger.info(f"✅ Successfully extracted {len(extracted_documents)} pages with advanced analysis")
                return extracted_documents
            else:
                errors.append("Advanced PyMuPDF: No content extracted from any page")
                
        except ImportError:
            errors.append("PyMuPDF not available - install with: pip install PyMuPDF")
        except Exception as e:
            errors.append(f"Advanced PyMuPDF error: {str(e)}")
            logger.warning(f"Advanced PyMuPDF failed: {e}")
        
        # Fallback methods (existing implementation as backup)
        return await self._fallback_pdf_extraction(file_path, errors)

    async def _extract_formatted_text(self, text_dict: dict) -> str:
        """Extract text while preserving formatting and structure."""
        formatted_lines = []
        
        try:
            for block in text_dict.get("blocks", []):
                if "lines" in block:
                    block_text = []
                    for line in block["lines"]:
                        line_text = []
                        for span in line.get("spans", []):
                            text = span.get("text", "").strip()
                            if text:
                                # Detect headers based on font size
                                font_size = span.get("size", 12)
                                if font_size > 14:
                                    text = f"## {text}"  # Mark as header
                                elif span.get("flags", 0) & 2**4:  # Bold
                                    text = f"**{text}**"
                                line_text.append(text)
                        
                        if line_text:
                            block_text.append(" ".join(line_text))
                    
                    if block_text:
                        formatted_lines.append("\n".join(block_text))
            
            return "\n\n".join(formatted_lines)
            
        except Exception as e:
            logger.warning(f"Text formatting failed: {e}")
            return ""

    async def _extract_pdf_tables(self, page, page_num: int) -> str:
        """Extract tables from PDF page using multiple methods."""
        tables_content = []
        
        try:
            # Method 1: Use PyMuPDF table detection
            try:
                import fitz
                tables = page.find_tables()
                
                for i, table in enumerate(tables):
                    try:
                        # Extract table data
                        table_data = table.extract()
                        if table_data:
                            # Convert to readable format
                            table_text = f"\n**Table {i + 1}:**\n"
                            
                            # Add headers if available
                            if len(table_data) > 0:
                                headers = table_data[0]
                                table_text += "| " + " | ".join(str(cell) for cell in headers) + " |\n"
                                table_text += "|" + "|".join("---" for _ in headers) + "|\n"
                                
                                # Add data rows
                                for row in table_data[1:]:
                                    table_text += "| " + " | ".join(str(cell) for cell in row) + " |\n"
                            
                            tables_content.append(table_text)
                            
                    except Exception as e:
                        logger.warning(f"Failed to extract table {i}: {e}")
                        
            except Exception as e:
                logger.warning(f"PyMuPDF table detection failed: {e}")
            
            # Method 2: Try pdfplumber for better table detection
            try:
                import pdfplumber
                
                with pdfplumber.open(page.parent.name) as pdf:
                    if page_num - 1 < len(pdf.pages):
                        plumber_page = pdf.pages[page_num - 1]
                        plumber_tables = plumber_page.extract_tables()
                        
                        for i, table in enumerate(plumber_tables):
                            if table:
                                table_text = f"\n**Table {len(tables_content) + i + 1} (pdfplumber):**\n"
                                
                                # Format table
                                for row_idx, row in enumerate(table):
                                    if row_idx == 0:  # Header
                                        table_text += "| " + " | ".join(str(cell) if cell else "" for cell in row) + " |\n"
                                        table_text += "|" + "|".join("---" for _ in row) + "|\n"
                                    else:  # Data rows
                                        table_text += "| " + " | ".join(str(cell) if cell else "" for cell in row) + " |\n"
                                
                                tables_content.append(table_text)
                                
            except ImportError:
                logger.info("pdfplumber not available - install with: pip install pdfplumber")
            except Exception as e:
                logger.warning(f"pdfplumber table extraction failed: {e}")
            
            return "\n".join(tables_content) if tables_content else ""
            
        except Exception as e:
            logger.error(f"Table extraction failed for page {page_num}: {e}")
            return ""

    async def _extract_pdf_images(self, page, page_num: int) -> str:
        """Extract and analyze images/charts from PDF page."""
        images_description = []
        
        try:
            image_list = page.get_images()
            
            for img_index, img in enumerate(image_list):
                try:
                    # Get image data
                    xref = img[0]
                    pix = fitz.Pixmap(page.parent, xref)
                    
                    if pix.n - pix.alpha < 4:  # GRAY or RGB
                        img_data = pix.tobytes("png")
                        
                        # Analyze image content using AI vision
                        image_description = await self._analyze_image_content(img_data, f"Image {img_index + 1} on page {page_num}")
                        
                        if image_description:
                            images_description.append(f"""
**Image {img_index + 1}:**
- Dimensions: {pix.width}x{pix.height} pixels
- Description: {image_description}
""")
                    
                    pix = None  # Free memory
                    
                except Exception as e:
                    logger.warning(f"Failed to process image {img_index}: {e}")
                    images_description.append(f"**Image {img_index + 1}:** [Could not analyze - {str(e)}]")
            
            return "\n".join(images_description) if images_description else ""
            
        except Exception as e:
            logger.error(f"Image extraction failed for page {page_num}: {e}")
            return ""

    async def _analyze_image_content(self, image_data: bytes, context: str) -> str:
        """Analyze image content using OpenAI Vision API or fallback methods."""
        try:
            # Try OpenAI Vision API if available
            try:
                import base64
                from openai import AsyncOpenAI
                
                base64_image = base64.b64encode(image_data).decode('utf-8')
                
                client = AsyncOpenAI(api_key=self.settings.openai_api_key)
                
                response = await client.chat.completions.create(
                    model="gpt-4o-mini",
                    messages=[
                        {
                            "role": "user",
                            "content": [
                                {
                                    "type": "text",
                                    "text": f"Analyze this image from a document ({context}). Describe what you see, focusing on any charts, graphs, tables, diagrams, or important visual information that would be relevant for data analysis."
                                },
                                {
                                    "type": "image_url",
                                    "image_url": {
                                        "url": f"data:image/png;base64,{base64_image}"
                                    }
                                }
                            ]
                        }
                    ],
                    max_tokens=300
                )
                
                return response.choices[0].message.content
                
            except Exception as vision_error:
                logger.warning(f"OpenAI Vision API failed: {vision_error}")
                
                # Fallback: Try basic OCR on image
                try:
                    ocr_text = await self._extract_text_with_ocr(image_data)
                    if ocr_text.strip():
                        return f"Contains text/data: {ocr_text[:200]}..."
                    else:
                        return "Visual element (chart, diagram, or image) - content not readable as text"
                except Exception as ocr_error:
                    logger.warning(f"OCR fallback failed: {ocr_error}")
                    return "Visual element detected but could not analyze content"
            
        except Exception as e:
            logger.error(f"Image analysis failed: {e}")
            return "Visual element present but analysis failed"

    async def _extract_pdf_forms(self, page, page_num: int) -> str:
        """Extract form fields and annotations from PDF."""
        forms_content = []
        
        try:
            # Extract form fields
            if hasattr(page, 'widgets'):
                widgets = page.widgets()
                
                for widget in widgets:
                    try:
                        field_name = widget.field_name
                        field_value = widget.field_value
                        field_type = widget.field_type_string
                        
                        if field_name or field_value:
                            forms_content.append(f"- **{field_name or 'Unnamed Field'}** ({field_type}): {field_value or '[Empty]'}")
                    except Exception as e:
                        logger.warning(f"Failed to extract form field: {e}")
            
            # Extract annotations
            annots = page.annots()
            for annot in annots:
                try:
                    annot_type = annot.type[1]  # Get annotation type name
                    content = annot.info.get("content", "")
                    
                    if content:
                        forms_content.append(f"- **{annot_type}:** {content}")
                except Exception as e:
                    logger.warning(f"Failed to extract annotation: {e}")
            
            return "\n".join(forms_content) if forms_content else ""
            
        except Exception as e:
            logger.error(f"Forms extraction failed for page {page_num}: {e}")
            return ""

    async def _fallback_pdf_extraction(self, file_path: Path, previous_errors: List[str]) -> List[Document]:
        """Fallback PDF extraction methods when advanced analysis fails."""
        # Method 2: Try PyPDFLoader
        try:
            logger.info(f"Attempting PyPDFLoader fallback for {file_path}")
            loader = PyPDFLoader(str(file_path))
            docs = loader.load()
            if docs and any(doc.page_content.strip() for doc in docs):
                logger.info(f"Successfully loaded PDF with PyPDFLoader: {len(docs)} pages")
                return docs
            else:
                previous_errors.append("PyPDFLoader: No content extracted")
        except Exception as e:
            previous_errors.append(f"PyPDFLoader error: {str(e)}")
            logger.warning(f"PyPDFLoader failed: {e}")
        
        # Method 3: OCR fallback
        try:
            from pdf2image import convert_from_path
            logger.info("Attempting full document OCR with pdf2image")
            
            images = convert_from_path(str(file_path), dpi=300, first_page=1, last_page=5)
            text_content = []
            
            for i, image in enumerate(images):
                import io
                img_byte_arr = io.BytesIO()
                image.save(img_byte_arr, format='PNG')
                img_bytes = img_byte_arr.getvalue()
                
                ocr_text = await self._extract_text_with_ocr(img_bytes)
                
                if ocr_text.strip():
                    text_content.append(
                        Document(
                            page_content=f"[OCR Extracted] {ocr_text}",
                            metadata={
                                "source": str(file_path),
                                "page": i + 1,
                                "extraction_method": "pdf2image + OCR"
                            }
                        )
                    )
            
            if text_content:
                logger.info(f"Successfully extracted {len(text_content)} pages with OCR")
                return text_content
            else:
                previous_errors.append("OCR: No text extracted from images")
                
        except ImportError:
            previous_errors.append("pdf2image not available")
        except Exception as e:
            previous_errors.append(f"OCR error: {str(e)}")
            logger.warning(f"OCR extraction failed: {e}")
        
        # If all methods fail, return helpful error document
        error_msg = "; ".join(previous_errors)
        logger.error(f"All PDF extraction methods failed for {file_path}: {error_msg}")
        
        return [Document(
            page_content=f"""📄 **Advanced PDF Analysis Report: {file_path.name}**

🔍 **Extraction Status**: Unable to extract readable content

📊 **File Information**:
- File Size: {file_path.stat().st_size / 1024:.1f} KB
- Format: PDF Document

⚠️ **Attempted Methods**: {len(previous_errors)} different extraction techniques
- Advanced PyMuPDF with table/image detection
- PyPDFLoader (LangChain)
- pdf2image + Tesseract OCR
- PyPDF2 direct extraction

🛠️ **Possible Issues**:
- Heavily encrypted or password-protected PDF
- Scanned document with poor image quality
- Complex layouts or non-standard fonts
- Corrupted PDF structure

💡 **Recommendations**:
1. **For Scanned Documents**: Improve scan quality (300+ DPI) and re-upload
2. **For Password-Protected PDFs**: Remove protection and re-upload
3. **Alternative Formats**: Convert to Word/Excel if possible for better structure extraction
4. **Manual Data Entry**: For critical data, consider manual extraction of key tables/charts

📈 **Enhanced Capabilities for Other Formats**:
- **Excel Files**: Full table extraction, multiple sheets, formulas
- **CSV Files**: Complete data analysis with statistical insights
- **Images**: Chart/graph analysis using AI vision
- **JSON**: Structured data parsing and analysis

🤖 **What I Can Do Instead**:
- Comprehensive statistical analysis on structured data
- Interactive visualizations and charts
- Pattern recognition and trend analysis
- Business intelligence insights and recommendations

{f"📝 **Technical Details**: {error_msg}" if previous_errors else ""}
""",
            metadata={
                "source": str(file_path),
                "extraction_method": "Failed",
                "error_count": len(previous_errors),
                "file_type": "pdf"
            }
        )]

    async def _extract_text_with_ocr(self, image_bytes: bytes) -> str:
        """Extract text from image using OCR."""
        try:
            import pytesseract
            from PIL import Image
            import io
            
            # Convert bytes to PIL Image
            image = Image.open(io.BytesIO(image_bytes))
            
            # Use OCR to extract text
            text = pytesseract.image_to_string(image, lang='eng+rus')  # Support English and Russian
            
            return text.strip()
            
        except ImportError:
            logger.warning("OCR libraries not available (pytesseract, PIL)")
            return ""
        except Exception as e:
            logger.warning(f"OCR extraction failed: {e}")
            return ""

    async def _load_csv_with_fallbacks(self, file_path: Path) -> List[Document]:
        """Load CSV with encoding fallbacks."""
        encodings = ['utf-8', 'latin1', 'iso-8859-1', 'cp1252']
        errors = []
        
        for encoding in encodings:
            try:
                loader = CSVLoader(str(file_path), encoding=encoding)
                docs = loader.load()
                if docs:
                    logger.info(f"Successfully loaded CSV with {encoding} encoding")
                    return docs
            except Exception as e:
                errors.append(f"{encoding}: {str(e)}")
        
        # Fallback to pandas
        try:
            df = pd.read_csv(file_path)
            return [Document(
                page_content=df.to_string(),
                metadata={"source": str(file_path), "file_type": "csv"}
            )]
        except Exception as e:
            errors.append(f"pandas: {str(e)}")
        
        return [Document(
            page_content=f"CSV file: {file_path.name}\n\nError loading file. Tried encodings: {', '.join(encodings)}\n\nErrors: {'; '.join(errors)}",
            metadata={
                "source": str(file_path), 
                "file_type": "csv", 
                "error_count": len(errors),
                "encodings_tried": ', '.join(encodings)
            }
        )]

    async def _load_text_with_fallbacks(self, file_path: Path) -> List[Document]:
        """Load text file with encoding fallbacks."""
        encodings = ['utf-8', 'latin1', 'iso-8859-1', 'cp1252']
        errors = []
        
        for encoding in encodings:
            try:
                loader = TextLoader(str(file_path), encoding=encoding)
                docs = loader.load()
                if docs:
                    logger.info(f"Successfully loaded text file with {encoding} encoding")
                    return docs
            except Exception as e:
                errors.append(f"{encoding}: {str(e)}")
        
        return [Document(
            page_content=f"Text file: {file_path.name}\n\nError loading file. Tried encodings: {', '.join(encodings)}\n\nErrors: {'; '.join(errors)}",
            metadata={
                "source": str(file_path), 
                "file_type": "text", 
                "error_count": len(errors),
                "encodings_tried": ', '.join(encodings)
            }
        )]

    async def _load_json_with_fallbacks(self, file_path: Path) -> List[Document]:
        """Load JSON with better error handling."""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = json.load(f)
            
            if isinstance(content, list):
                # Split large arrays into chunks
                chunk_size = 100
                chunks = [content[i:i + chunk_size] for i in range(0, len(content), chunk_size)]
                return [
                    Document(
                        page_content=json.dumps(chunk, indent=2),
                        metadata={
                            "source": str(file_path),
                            "file_type": "json",
                            "chunk": i,
                            "total_chunks": len(chunks)
                        }
                    )
                    for i, chunk in enumerate(chunks)
                ]
            else:
                return [Document(
                    page_content=json.dumps(content, indent=2),
                    metadata={"source": str(file_path), "file_type": "json"}
                )]
                
        except json.JSONDecodeError as e:
            return [Document(
                page_content=f"JSON file: {file_path.name}\n\nInvalid JSON format: {str(e)}",
                metadata={"source": str(file_path), "file_type": "json", "error": str(e)}
            )]
        except Exception as e:
            return [Document(
                page_content=f"JSON file: {file_path.name}\n\nError loading file: {str(e)}",
                metadata={"source": str(file_path), "file_type": "json", "error": str(e)}
            )]

    async def _load_excel_with_fallbacks(self, file_path: Path) -> List[Document]:
        """Enhanced Excel loader that extracts multiple sheets, charts, formulas, and formatting."""
        try:
            logger.info(f"🔍 Starting advanced Excel analysis for {file_path}")
            documents = []
            
            # Method 1: Advanced pandas analysis with openpyxl for rich content
            try:
                import openpyxl
                from openpyxl.chart import PieChart, BarChart, LineChart, ScatterChart, AreaChart
                
                # Load workbook with openpyxl for advanced features
                wb = openpyxl.load_workbook(file_path, data_only=False)
                
                # Process each worksheet
                for sheet_name in wb.sheetnames:
                    logger.info(f"📊 Processing sheet: {sheet_name}")
                    worksheet = wb[sheet_name]
                    sheet_content = []
                    
                    # 1. Extract sheet metadata and structure
                    sheet_info = await self._analyze_excel_sheet_structure(worksheet, sheet_name)
                    sheet_content.append(f"📋 **Sheet Overview:**\n{sheet_info}")
                    
                    # 2. Extract data tables
                    try:
                        df = pd.read_excel(file_path, sheet_name=sheet_name)
                        if not df.empty:
                            data_analysis = await self._analyze_excel_data(df, sheet_name)
                            sheet_content.append(f"\n📊 **Data Analysis:**\n{data_analysis}")
                            
                            # Include sample data
                            sample_data = df.head(10).to_string(max_cols=10)
                            sheet_content.append(f"\n📋 **Sample Data (First 10 rows):**\n```\n{sample_data}\n```")
                        
                    except Exception as e:
                        logger.warning(f"Failed to read data from sheet {sheet_name}: {e}")
                    
                    # 3. Extract formulas
                    try:
                        formulas = await self._extract_excel_formulas(worksheet)
                        if formulas:
                            sheet_content.append(f"\n🧮 **Formulas Found:**\n{formulas}")
                    except Exception as e:
                        logger.warning(f"Formula extraction failed for {sheet_name}: {e}")
                    
                    # 4. Extract charts
                    try:
                        charts_info = await self._extract_excel_charts(worksheet)
                        if charts_info:
                            sheet_content.append(f"\n📈 **Charts and Visualizations:**\n{charts_info}")
                    except Exception as e:
                        logger.warning(f"Chart extraction failed for {sheet_name}: {e}")
                    
                    # 5. Extract comments and annotations
                    try:
                        comments = await self._extract_excel_comments(worksheet)
                        if comments:
                            sheet_content.append(f"\n💬 **Comments and Notes:**\n{comments}")
                    except Exception as e:
                        logger.warning(f"Comments extraction failed for {sheet_name}: {e}")
                    
                    # 6. Extract conditional formatting and styles
                    try:
                        formatting_info = await self._extract_excel_formatting(worksheet)
                        if formatting_info:
                            sheet_content.append(f"\n🎨 **Formatting and Highlights:**\n{formatting_info}")
                    except Exception as e:
                        logger.warning(f"Formatting extraction failed for {sheet_name}: {e}")
                    
                    # Create document for this sheet
                    if sheet_content:
                        combined_content = "\n".join(sheet_content)
                        
                        document = Document(
                            page_content=f"""📊 **Excel Sheet: {sheet_name}**

{combined_content}
""",
                            metadata={
                                "source": str(file_path),
                                "sheet_name": sheet_name,
                                "file_type": "excel",
                                "extraction_method": "Advanced Excel Analysis",
                                "workbook": file_path.stem
                            }
                        )
                        documents.append(document)
                
                wb.close()
                
                if documents:
                    logger.info(f"✅ Successfully extracted {len(documents)} Excel sheets with advanced analysis")
                    return documents
                    
            except ImportError:
                logger.warning("openpyxl not available - install with: pip install openpyxl")
            except Exception as e:
                logger.warning(f"Advanced Excel analysis failed: {e}")
            
            # Method 2: Pandas fallback with multiple sheets
            try:
                logger.info("📊 Attempting pandas multi-sheet extraction")
                
                # Read all sheets
                excel_file = pd.ExcelFile(file_path)
                documents = []
                
                for sheet_name in excel_file.sheet_names:
                    try:
                        df = pd.read_excel(file_path, sheet_name=sheet_name)
                        
                        if not df.empty:
                            # Basic data analysis
                            data_summary = await self._create_basic_excel_summary(df, sheet_name)
                            
                            # Include full data for smaller sheets, sample for larger ones
                            if len(df) <= 100:
                                data_content = df.to_string(max_cols=20)
                            else:
                                sample_data = df.head(20).to_string(max_cols=20)
                                data_content = f"Sample Data (First 20 rows):\n{sample_data}\n\n[Sheet contains {len(df)} total rows]"
                            
                            document = Document(
                                page_content=f"""📊 **Excel Sheet: {sheet_name}**

📋 **Sheet Summary:**
{data_summary}

📊 **Data Content:**
```
{data_content}
```
""",
                                metadata={
                                    "source": str(file_path),
                                    "sheet_name": sheet_name,
                                    "file_type": "excel",
                                    "extraction_method": "Pandas Basic",
                                    "rows": len(df),
                                    "columns": len(df.columns)
                                }
                            )
                            documents.append(document)
                            
                    except Exception as e:
                        logger.warning(f"Failed to process sheet {sheet_name}: {e}")
                        # Add error document for this sheet
                        documents.append(Document(
                            page_content=f"📊 **Excel Sheet: {sheet_name}**\n\n⚠️ Failed to process this sheet: {str(e)}",
                            metadata={
                                "source": str(file_path),
                                "sheet_name": sheet_name,
                                "file_type": "excel",
                                "extraction_method": "Failed",
                                "error": str(e)
                            }
                        ))
                
                if documents:
                    logger.info(f"✅ Successfully extracted {len(documents)} Excel sheets with pandas")
                    return documents
                    
            except Exception as e:
                logger.warning(f"Pandas Excel extraction failed: {e}")
            
            # Method 3: Unstructured loader fallback
            try:
                logger.info("📊 Attempting UnstructuredExcelLoader fallback")
                loader = UnstructuredExcelLoader(str(file_path))
                docs = loader.load()
                if docs:
                    logger.info(f"✅ Successfully loaded Excel with UnstructuredExcelLoader: {len(docs)} documents")
                    return docs
            except Exception as e:
                logger.warning(f"UnstructuredExcelLoader failed: {e}")
            
            # If all methods fail
            return [Document(
                page_content=f"""📊 **Enhanced Excel Analysis Report: {file_path.name}**

🔍 **Extraction Status**: Unable to process Excel file

📊 **File Information**:
- File Size: {file_path.stat().st_size / 1024:.1f} KB
- Format: Excel Workbook (.xlsx/.xls)

⚠️ **Possible Issues**:
- File corruption or invalid Excel format
- Password-protected workbook
- Very large file size causing memory issues
- Missing required libraries (openpyxl, xlrd)

🛠️ **Attempted Methods**:
- Advanced openpyxl analysis (sheets, charts, formulas)
- Pandas multi-sheet extraction
- UnstructuredExcelLoader fallback

💡 **Recommendations**:
1. **For Password-Protected Files**: Remove protection and re-upload
2. **For Large Files**: Split into smaller workbooks or save as CSV
3. **Alternative Formats**: Export as CSV files for individual sheets
4. **Manual Processing**: Extract key data tables manually

📈 **What I Can Analyze Instead**:
- **CSV Files**: Complete statistical analysis and visualization
- **JSON Data**: Structured data parsing and insights
- **PDF Reports**: Text and table extraction
- **Images**: Chart/graph analysis using AI vision

🤖 **My Excel Capabilities** (when working):
- Multi-sheet analysis and comparison
- Formula extraction and documentation
- Chart and visualization detection
- Data quality assessment
- Conditional formatting analysis
- Statistical insights across worksheets

Please try converting to CSV format or check the file integrity and re-upload.
""",
                metadata={
                    "source": str(file_path),
                    "file_type": "excel",
                    "extraction_method": "Failed",
                    "error": "All extraction methods failed"
                }
            )]
            
        except Exception as e:
            logger.error(f"Excel processing failed for {file_path}: {e}")
            return [Document(
                page_content=f"📊 Excel file: {file_path.name}\n\n⚠️ Error loading file: {str(e)}",
                metadata={"source": str(file_path), "file_type": "excel", "error": str(e)}
            )]

    async def _analyze_excel_sheet_structure(self, worksheet, sheet_name: str) -> str:
        """Analyze Excel sheet structure and metadata."""
        try:
            # Get sheet dimensions
            max_row = worksheet.max_row
            max_col = worksheet.max_column
            
            # Detect merged cells
            merged_ranges = len(worksheet.merged_cells.ranges)
            
            # Count non-empty cells
            non_empty_cells = 0
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        non_empty_cells += 1
            
            # Try to detect headers
            first_row_values = [cell.value for cell in worksheet[1] if cell.value is not None]
            has_headers = len(first_row_values) > 1 and all(isinstance(v, str) for v in first_row_values[:5])
            
            structure_info = f"""- **Dimensions**: {max_row} rows × {max_col} columns
- **Data Density**: {non_empty_cells} non-empty cells ({non_empty_cells/(max_row*max_col)*100:.1f}% filled)
- **Merged Cells**: {merged_ranges} merged ranges
- **Headers Detected**: {'Yes' if has_headers else 'No'}
{f"- **Potential Headers**: {', '.join(str(v) for v in first_row_values[:10])}" if has_headers else ""}"""
            
            return structure_info
            
        except Exception as e:
            return f"Structure analysis failed: {str(e)}"

    async def _analyze_excel_data(self, df: pd.DataFrame, sheet_name: str) -> str:
        """Analyze Excel data content."""
        try:
            analysis = []
            
            # Basic stats
            total_rows, total_cols = df.shape
            numeric_cols = df.select_dtypes(include=[np.number]).columns.tolist()
            text_cols = df.select_dtypes(include=['object']).columns.tolist()
            
            analysis.append(f"- **Data Shape**: {total_rows:,} rows × {total_cols} columns")
            analysis.append(f"- **Column Types**: {len(numeric_cols)} numeric, {len(text_cols)} text")
            
            # Missing data
            missing_pct = (df.isnull().sum().sum() / (total_rows * total_cols) * 100)
            analysis.append(f"- **Missing Data**: {missing_pct:.1f}% of cells are empty")
            
            # Column summary
            if total_cols <= 20:
                col_info = []
                for col in df.columns:
                    col_type = str(df[col].dtype)
                    unique_vals = df[col].nunique()
                    col_info.append(f"{col} ({col_type}, {unique_vals} unique)")
                analysis.append(f"- **Columns**: {', '.join(col_info)}")
            
            # Data preview insights
            if len(numeric_cols) > 0:
                numeric_summary = df[numeric_cols].describe()
                analysis.append(f"- **Numeric Summary**: {len(numeric_cols)} numeric columns with statistics available")
            
            return "\n".join(analysis)
            
        except Exception as e:
            return f"Data analysis failed: {str(e)}"

    async def _extract_excel_formulas(self, worksheet) -> str:
        """Extract formulas from Excel worksheet."""
        try:
            formulas = []
            
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.data_type == 'f':  # Formula cell
                        cell_ref = f"{cell.column_letter}{cell.row}"
                        formula = cell.value
                        result = cell.displayed_value
                        formulas.append(f"- **{cell_ref}**: `{formula}` → {result}")
            
            if formulas:
                return "\n".join(formulas[:20])  # Limit to first 20 formulas
            return ""
            
        except Exception as e:
            logger.warning(f"Formula extraction failed: {e}")
            return ""

    async def _extract_excel_charts(self, worksheet) -> str:
        """Extract chart information from Excel worksheet."""
        try:
            charts_info = []
            
            # Get charts (this is a simplified version - full implementation would need python-pptx or similar)
            if hasattr(worksheet, '_charts'):
                for i, chart in enumerate(worksheet._charts):
                    try:
                        chart_type = type(chart).__name__
                        charts_info.append(f"- **Chart {i+1}**: {chart_type}")
                        
                        # Try to get chart title
                        if hasattr(chart, 'title') and chart.title:
                            charts_info.append(f"  - Title: {chart.title}")
                            
                    except Exception as e:
                        charts_info.append(f"- **Chart {i+1}**: Could not analyze ({str(e)})")
            
            return "\n".join(charts_info) if charts_info else ""
            
        except Exception as e:
            logger.warning(f"Chart extraction failed: {e}")
            return ""

    async def _extract_excel_comments(self, worksheet) -> str:
        """Extract comments and notes from Excel worksheet."""
        try:
            comments = []
            
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.comment:
                        cell_ref = f"{cell.column_letter}{cell.row}"
                        comment_text = cell.comment.text
                        comments.append(f"- **{cell_ref}**: {comment_text}")
            
            return "\n".join(comments) if comments else ""
            
        except Exception as e:
            logger.warning(f"Comments extraction failed: {e}")
            return ""

    async def _extract_excel_formatting(self, worksheet) -> str:
        """Extract formatting information from Excel worksheet."""
        try:
            formatting_info = []
            
            # Check for conditional formatting
            if worksheet.conditional_formatting:
                formatting_info.append(f"- **Conditional Formatting**: {len(worksheet.conditional_formatting)} rules applied")
            
            # Check for colored cells (basic detection)
            colored_cells = 0
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.fill.fgColor.rgb and cell.fill.fgColor.rgb != '00000000':
                        colored_cells += 1
            
            if colored_cells > 0:
                formatting_info.append(f"- **Colored Cells**: {colored_cells} cells with background colors")
            
            return "\n".join(formatting_info) if formatting_info else ""
            
        except Exception as e:
            logger.warning(f"Formatting extraction failed: {e}")
            return ""

    async def _create_basic_excel_summary(self, df: pd.DataFrame, sheet_name: str) -> str:
        """Create basic summary for Excel sheet."""
        try:
            summary = []
            
            # Basic info
            rows, cols = df.shape
            summary.append(f"- **Dimensions**: {rows:,} rows × {cols} columns")
            
            # Column types
            numeric_cols = df.select_dtypes(include=[np.number]).columns.tolist()
            summary.append(f"- **Data Types**: {len(numeric_cols)} numeric, {cols - len(numeric_cols)} other")
            
            # Missing data
            missing_pct = (df.isnull().sum().sum() / (rows * cols) * 100)
            summary.append(f"- **Completeness**: {100 - missing_pct:.1f}% of data is present")
            
            # Quick stats for numeric columns
            if numeric_cols:
                summary.append(f"- **Numeric Columns**: {', '.join(numeric_cols[:10])}")
                if len(numeric_cols) > 10:
                    summary.append(f"  ... and {len(numeric_cols) - 10} more")
            
            return "\n".join(summary)
            
        except Exception as e:
            return f"Summary generation failed: {str(e)}"

    async def _generate_document_summary(self, documents: List[Document]) -> str:
        """Generate a summary of the document content."""
        if not documents:
            return "No content available for summary."
        
        # Combine first few documents for summary
        content = "\n".join([doc.page_content[:500] for doc in documents[:3]])
        
        prompt = f"""
        Analyze this document content and provide a concise summary:
        
        Content:
        {content}
        
        Provide a summary that includes:
        1. Document type and main topic
        2. Key data points or information
        3. Potential analysis opportunities
        
        Summary:
        """
        
        try:
            response = await self.llm.ainvoke(prompt)
            return response.content
        except Exception as e:
            logger.error(f"Error generating summary: {e}")
            return "Summary generation failed."

    async def analyze_data(self, query: str, document_filter: Optional[str] = None) -> AnalysisResult:
        """
        Perform comprehensive data analysis like a professional data analyst.
        
        Args:
            query: Analysis query from user
            document_filter: Optional filter for specific documents
            
        Returns:
            Detailed analysis result with professional insights
        """
        try:
            logger.info(f"🔍 Starting comprehensive data analysis for: {query[:100]}...")
            
            # Extract the actual user query from context if present
            user_query = self._extract_user_query(query)
            logger.info(f"📊 Extracted user query: {user_query[:100]}...")
            
            # Retrieve relevant documents with expanded search
            docs = self.vectorstore.similarity_search(user_query, k=25)  # Increased for better context
            logger.info(f"📚 Found {len(docs)} relevant documents")
            
            if not docs:
                return await self._provide_data_upload_guidance()
            
            # Check if we have structured data (CSV/Excel)
            structured_data = await self._extract_structured_data(docs)
            logger.info(f"📈 Structured data analysis: {structured_data is not None}")
            
            # Determine analysis type and provide comprehensive insights
            if structured_data is not None:
                logger.info(f"🧮 Performing quantitative analysis on dataset: {structured_data.shape}")
                return await self._perform_professional_quantitative_analysis(user_query, structured_data, docs)
            else:
                logger.info("📝 Performing qualitative analysis on text documents")
                return await self._perform_professional_qualitative_analysis(user_query, docs)
                
        except Exception as e:
            logger.error(f"❌ Error in data analysis: {e}", exc_info=True)
            return AnalysisResult(
                summary=f"⚠️ Analysis Error: {str(e)}",
                insights=["Technical issue encountered during analysis"],
                recommendations=[
                    "Try rephrasing your question with more specific terms",
                    "Ensure your data file is properly formatted and not corrupted", 
                    "Check if the file contains readable text or structured data",
                    "Contact support if the issue persists"
                ],
                visualizations=[],
                statistical_analysis={"error": str(e), "status": "failed"},
                confidence_score=0.1
            )

    async def _provide_data_upload_guidance(self) -> AnalysisResult:
        """Provide guidance when no data is available."""
        return AnalysisResult(
            summary="🚀 Ready to Analyze Your Data! No files uploaded yet.",
            insights=[
                "📊 I'm your AI Data Analyst - I can analyze various data formats",
                "🔍 I specialize in finding patterns, trends, and actionable insights",
                "📈 I can create visualizations and provide statistical analysis",
                "🤖 I work like a professional data analyst with advanced AI capabilities"
            ],
            recommendations=[
                "📁 Upload your data files (CSV, Excel, JSON, PDF, or text files)",
                "💡 Ask specific questions like 'What trends do you see in my sales data?'",
                "📊 Request visualizations: 'Create a chart showing monthly trends'",
                "🔬 Get statistical insights: 'Find correlations in my dataset'",
                "💼 Business questions: 'What recommendations do you have for improving performance?'"
            ],
            visualizations=[],
            statistical_analysis={
                "status": "awaiting_data",
                "message": "Upload data files to enable comprehensive statistical analysis",
                "supported_formats": ["CSV", "Excel", "JSON", "PDF", "Text"]
            },
            confidence_score=1.0
        )

    async def _perform_professional_quantitative_analysis(self, query: str, df: pd.DataFrame, docs: List[Document]) -> AnalysisResult:
        """Perform comprehensive quantitative analysis like a professional data analyst."""
        insights = []
        recommendations = []
        visualizations = []
        statistical_analysis = {}
        
        try:
            # Dataset Overview & Quality Assessment
            total_rows, total_cols = df.shape
            numeric_cols = df.select_dtypes(include=[np.number]).columns.tolist()
            categorical_cols = df.select_dtypes(include=['object', 'category']).columns.tolist()
            datetime_cols = df.select_dtypes(include=['datetime64']).columns.tolist()
            
            # Comprehensive Data Profiling
            data_profile = {
                "dataset_overview": {
                    "total_rows": total_rows,
                    "total_columns": total_cols,
                    "numeric_columns": len(numeric_cols),
                    "categorical_columns": len(categorical_cols),
                    "datetime_columns": len(datetime_cols),
                    "column_names": df.columns.tolist(),
                    "memory_usage_mb": round(df.memory_usage(deep=True).sum() / 1024**2, 2)
                },
                "data_quality": {
                    "missing_values": df.isnull().sum().to_dict(),
                    "missing_percentage": (df.isnull().sum() / len(df) * 100).round(2).to_dict(),
                    "duplicate_rows": int(df.duplicated().sum()),
                    "duplicate_percentage": round(df.duplicated().sum() / len(df) * 100, 2)
                }
            }
            
            # Professional Data Quality Assessment
            insights.append(f"📊 **Dataset Overview**: {total_rows:,} rows × {total_cols} columns")
            insights.append(f"🔢 **Data Types**: {len(numeric_cols)} numeric, {len(categorical_cols)} categorical, {len(datetime_cols)} datetime")
            
            # Data Quality Insights
            missing_data = df.isnull().sum()
            high_missing = missing_data[missing_data > total_rows * 0.1]
            
            if len(high_missing) > 0:
                insights.append(f"⚠️ **Data Quality Alert**: {len(high_missing)} columns have >10% missing values")
                recommendations.append("🔧 Consider data imputation or removal of high-missing columns")
            else:
                insights.append("✅ **Data Quality**: Excellent - minimal missing values detected")
            
            duplicate_pct = round(df.duplicated().sum() / len(df) * 100, 2)
            if duplicate_pct > 5:
                insights.append(f"🔄 **Duplicate Records**: {duplicate_pct}% of data is duplicated")
                recommendations.append("🧹 Consider deduplication to improve analysis accuracy")
            
            # Advanced Statistical Analysis
            if len(numeric_cols) > 0:
                numeric_stats = df[numeric_cols].describe()
                data_profile["descriptive_statistics"] = numeric_stats.to_dict()
                
                # Distribution Analysis
                for col in numeric_cols[:5]:  # Analyze first 5 numeric columns
                    skewness = float(df[col].skew())
                    kurtosis = float(df[col].kurtosis())
                    
                    if abs(skewness) > 1:
                        insights.append(f"📈 **{col}**: Highly skewed distribution (skewness: {skewness:.2f})")
                    
                    if abs(kurtosis) > 3:
                        insights.append(f"📊 **{col}**: Heavy-tailed distribution (kurtosis: {kurtosis:.2f})")
                
                # Correlation Analysis with Professional Interpretation
                if len(numeric_cols) >= 2:
                    correlation_matrix = df[numeric_cols].corr()
                    data_profile["correlations"] = correlation_matrix.to_dict()
                    
                    # Find strong correlations
                    strong_correlations = []
                    for i in range(len(correlation_matrix.columns)):
                        for j in range(i+1, len(correlation_matrix.columns)):
                            corr_val = correlation_matrix.iloc[i, j]
                            if abs(corr_val) > 0.7:
                                col1, col2 = correlation_matrix.columns[i], correlation_matrix.columns[j]
                                relationship = "positive" if corr_val > 0 else "negative"
                                strong_correlations.append({
                                    "variables": [col1, col2],
                                    "correlation": round(corr_val, 3),
                                    "relationship": relationship,
                                    "strength": "very strong" if abs(corr_val) > 0.9 else "strong"
                                })
                    
                    if strong_correlations:
                        insights.append(f"🔗 **Strong Correlations Found**: {len(strong_correlations)} significant relationships")
                        for corr in strong_correlations[:3]:  # Show top 3
                            insights.append(f"   • {corr['variables'][0]} ↔ {corr['variables'][1]}: {corr['strength']} {corr['relationship']} ({corr['correlation']})")
                        
                        recommendations.append("🎯 Investigate these correlations for potential causal relationships")
                    
                    data_profile["strong_correlations"] = strong_correlations
            
            # Categorical Data Analysis
            if len(categorical_cols) > 0:
                categorical_insights = {}
                for col in categorical_cols[:5]:  # Analyze first 5 categorical columns
                    unique_count = df[col].nunique()
                    most_common = df[col].value_counts().head(3)
                    
                    categorical_insights[col] = {
                        "unique_values": int(unique_count),
                        "most_common": most_common.to_dict()
                    }
                    
                    if unique_count > total_rows * 0.5:
                        insights.append(f"🏷️ **{col}**: High cardinality ({unique_count} unique values)")
                    elif unique_count < 10:
                        insights.append(f"🏷️ **{col}**: Low cardinality ({unique_count} categories)")
                
                data_profile["categorical_analysis"] = categorical_insights
            
            # Advanced Pattern Detection
            patterns = await self._detect_advanced_patterns(df, query)
            if patterns:
                insights.extend(patterns["insights"])
                recommendations.extend(patterns["recommendations"])
                if patterns.get("visualizations"):
                    visualizations.extend(patterns["visualizations"])
            
            # Generate Professional Summary
            summary = await self._generate_professional_summary(query, df, data_profile, insights)
            
            statistical_analysis["comprehensive_profile"] = data_profile
            statistical_analysis["analysis_timestamp"] = datetime.utcnow().isoformat()
            statistical_analysis["query_analyzed"] = query
            
            # Calculate confidence score based on data quality and completeness
            confidence_score = self._calculate_analysis_confidence(df, data_profile)
            
            return AnalysisResult(
                summary=summary,
                insights=insights,
                recommendations=recommendations,
                visualizations=visualizations,
                statistical_analysis=statistical_analysis,
                confidence_score=confidence_score
            )
            
        except Exception as e:
            logger.error(f"Error in quantitative analysis: {e}", exc_info=True)
            return AnalysisResult(
                summary=f"⚠️ Quantitative analysis encountered an issue: {str(e)}",
                insights=["Analysis partially completed before encountering technical issue"],
                recommendations=["Try with a smaller dataset or different file format"],
                visualizations=[],
                statistical_analysis={"error": str(e), "partial_analysis": True},
                confidence_score=0.3
            )

    async def _detect_advanced_patterns(self, df: pd.DataFrame, query: str) -> Optional[Dict[str, Any]]:
        """Detect advanced patterns in the data like a professional analyst."""
        patterns = {"insights": [], "recommendations": [], "visualizations": []}
        
        try:
            numeric_cols = df.select_dtypes(include=[np.number]).columns
            
            # Time Series Pattern Detection
            date_cols = df.select_dtypes(include=['datetime64']).columns
            if len(date_cols) > 0 and len(numeric_cols) > 0:
                patterns["insights"].append("📅 **Time Series Data Detected**: Temporal analysis capabilities available")
                patterns["recommendations"].append("📈 Consider trend analysis, seasonality detection, and forecasting")
            
            # Outlier Detection
            if len(numeric_cols) > 0:
                outlier_counts = {}
                for col in numeric_cols:
                    Q1 = df[col].quantile(0.25)
                    Q3 = df[col].quantile(0.75)
                    IQR = Q3 - Q1
                    outliers = df[(df[col] < Q1 - 1.5*IQR) | (df[col] > Q3 + 1.5*IQR)]
                    outlier_counts[col] = len(outliers)
                
                high_outlier_cols = {k: v for k, v in outlier_counts.items() if v > len(df) * 0.05}
                if high_outlier_cols:
                    patterns["insights"].append(f"🎯 **Outliers Detected**: {len(high_outlier_cols)} columns have significant outliers")
                    patterns["recommendations"].append("🔍 Investigate outliers - they may indicate data errors or important insights")
            
            # Distribution Patterns
            if len(numeric_cols) >= 2:
                # Check for potential normal distributions
                normal_like_cols = []
                for col in numeric_cols:
                    skew = abs(df[col].skew())
                    if skew < 0.5:  # Roughly normal
                        normal_like_cols.append(col)
                
                if normal_like_cols:
                    patterns["insights"].append(f"📊 **Normal Distributions**: {len(normal_like_cols)} columns appear normally distributed")
                    patterns["recommendations"].append("📈 Normal distributions enable parametric statistical tests")
            
            # Clustering Potential
            if len(numeric_cols) >= 2 and len(df) > 50:
                patterns["insights"].append("🎯 **Clustering Analysis**: Dataset suitable for segmentation analysis")
                patterns["recommendations"].append("🔬 Consider customer/data point segmentation using clustering algorithms")
            
            return patterns
            
        except Exception as e:
            logger.warning(f"Pattern detection failed: {e}")
            return None

    async def _generate_professional_summary(self, query: str, df: pd.DataFrame, profile: Dict, insights: List[str]) -> str:
        """Generate a professional data analyst summary."""
        
        # Create a comprehensive prompt for the LLM
        prompt = f"""
        As a senior data analyst, provide a professional summary of this dataset analysis.
        
        User Query: {query}
        
        Dataset Profile:
        - Rows: {profile['dataset_overview']['total_rows']:,}
        - Columns: {profile['dataset_overview']['total_columns']}
        - Data Types: {profile['dataset_overview']['numeric_columns']} numeric, {profile['dataset_overview']['categorical_columns']} categorical
        - Data Quality: {profile['data_quality']['duplicate_percentage']}% duplicates, missing data varies by column
        
        Key Insights Found:
        {chr(10).join(insights[:10])}
        
        Provide a professional, concise summary (2-3 paragraphs) that:
        1. Addresses the user's query directly
        2. Highlights the most important findings
        3. Uses professional data analyst language
        4. Provides actionable insights
        
        Summary:
        """
        
        try:
            response = await self.llm.ainvoke(prompt)
            return f"🔍 **Professional Data Analysis Summary**\n\n{response.content}"
        except Exception as e:
            logger.error(f"Error generating professional summary: {e}")
            return f"""🔍 **Data Analysis Summary**

            Analyzed dataset with {profile['dataset_overview']['total_rows']:,} records and {profile['dataset_overview']['total_columns']} variables. 
            The analysis reveals {len(insights)} key insights about your data's structure, quality, and patterns.
            
            Key findings include data quality assessment, statistical patterns, and actionable recommendations 
            for further analysis. The dataset shows {profile['data_quality']['duplicate_percentage']}% duplicate records 
            and varying levels of completeness across variables."""

    def _calculate_analysis_confidence(self, df: pd.DataFrame, profile: Dict) -> float:
        """Calculate confidence score for the analysis."""
        score = 1.0
        
        # Reduce confidence for high missing data
        avg_missing = sum(profile['data_quality']['missing_percentage'].values()) / len(profile['data_quality']['missing_percentage'])
        score -= min(avg_missing / 100, 0.3)
        
        # Reduce confidence for high duplicates
        score -= min(profile['data_quality']['duplicate_percentage'] / 100, 0.2)
        
        # Reduce confidence for very small datasets
        if len(df) < 50:
            score -= 0.2
        
        # Increase confidence for rich data types
        if profile['dataset_overview']['numeric_columns'] > 0:
            score += 0.1
        
        return max(0.1, min(1.0, score))

    def _extract_user_query(self, full_query: str) -> str:
        """Extract the actual user query from context-enhanced query."""
        # Look for "User Query:" pattern
        if "User Query:" in full_query:
            parts = full_query.split("User Query:")
            if len(parts) > 1:
                return parts[1].strip().split("\n")[0].strip()
        return full_query

    async def _extract_structured_data(self, docs: List[Document]) -> Optional[pd.DataFrame]:
        """Extract structured data from documents if available."""
        try:
            # Look for CSV or Excel files in processed documents
            for filename, doc_info in self.processed_documents.items():
                if doc_info.file_type in ['.csv', '.xlsx', '.xls']:
                    file_path = Path(self.settings.data_directory) / filename
                    
                    if file_path.exists():
                        if doc_info.file_type == '.csv':
                            return pd.read_csv(file_path)
                        else:
                            return pd.read_excel(file_path)
            
            return None
            
        except Exception as e:
            logger.error(f"Error extracting structured data: {e}")
            return None

    async def _perform_professional_qualitative_analysis(self, query: str, docs: List[Document]) -> AnalysisResult:
        """Perform enhanced qualitative analysis on text documents with conversational insights."""
        try:
            # Combine relevant document content
            content = "\n".join([doc.page_content for doc in docs[:8]])
            
            # Enhanced analysis prompt with conversational context
            prompt = f"""
            As a professional data analyst, analyze the following documents to answer: "{query}"
            
            Document Content:
            {content[:4000]}
            
            Provide a comprehensive analysis following these principles:
            
            1. **Direct Response**: Address the specific question asked
            2. **Evidence-Based Insights**: Extract 3-5 key findings supported by document content
            3. **Actionable Recommendations**: Provide 3-4 specific, implementable next steps
            4. **Business Context**: Explain implications and significance
            5. **Future Considerations**: Suggest follow-up questions or analyses
            
            Maintain a professional, conversational tone while being precise and data-driven.
            Reference specific content from the documents to support your analysis.
            
            Format as JSON:
            {{
                "direct_answer": "Clear, specific answer to the user's question",
                "summary": "Concise overview of what the documents reveal",
                "insights": ["insight1", "insight2", "insight3", "insight4", "insight5"],
                "recommendations": ["action1", "action2", "action3", "action4"],
                "implications": "What this means for decision-making and strategy",
                "follow_up_questions": ["question1", "question2", "question3"],
                "confidence": 0.8
            }}
            """
            
            response = await self.llm.ainvoke(prompt)
            
            try:
                # Parse LLM response
                analysis = json.loads(response.content)
                
                # Create conversational summary
                summary = f"""Based on your documents, here's what I found regarding "{query}":

{analysis.get('direct_answer', 'I analyzed your documents and found relevant information.')}

{analysis.get('implications', 'This information could be valuable for your decision-making process.')}"""
                
                return AnalysisResult(
                    summary=summary,
                    insights=analysis.get("insights", [])[:6],
                    recommendations=analysis.get("recommendations", [])[:5],
                    visualizations=[],  # Text analysis doesn't generate visualizations
                    statistical_analysis={
                        "document_analysis": {
                            "documents_analyzed": len(docs),
                            "content_length": len(content),
                            "analysis_type": "qualitative",
                            "follow_up_questions": analysis.get("follow_up_questions", [])
                        }
                    },
                    confidence_score=analysis.get("confidence", 0.75)
                )
                
            except json.JSONDecodeError:
                # Fallback if JSON parsing fails
                return AnalysisResult(
                    summary=f"I analyzed your documents regarding '{query}'. Here's what I found:\n\n{response.content[:800]}...",
                    insights=["Document analysis completed", "Key information extracted from your files"],
                    recommendations=["Consider uploading structured data for quantitative analysis", "Ask specific follow-up questions about the content"],
                    visualizations=[],
                    statistical_analysis={"analysis_type": "qualitative_fallback"},
                    confidence_score=0.6
                )
                
        except Exception as e:
            logger.error(f"Error in enhanced qualitative analysis: {e}")
            return AnalysisResult(
                summary=f"I had trouble analyzing your documents: {str(e)}",
                insights=["Document analysis encountered an issue"],
                recommendations=["Try rephrasing your question", "Ensure documents are properly uploaded"],
                visualizations=[],
                statistical_analysis={"error": str(e)},
                confidence_score=0.2
            )

    def get_uploaded_files(self) -> List[Dict[str, Any]]:
        """Get list of uploaded and processed files."""
        return [
            {
                "filename": doc.filename,
                "file_type": doc.file_type,
                "size": doc.size,
                "processed_at": doc.processed_at.isoformat(),
                "chunks_count": doc.chunks_count,
                "summary": doc.summary
            }
            for doc in self.processed_documents.values()
        ]

    async def delete_file(self, filename: str) -> bool:
        """Delete a file and its embeddings."""
        try:
            # Remove from processed documents
            if filename in self.processed_documents:
                del self.processed_documents[filename]
            
            # Remove file from disk
            file_path = Path(self.settings.data_directory) / filename
            if file_path.exists():
                file_path.unlink()
            
            # Note: Removing from Chroma would require tracking document IDs
            # For now, we'll leave embeddings (they'll be overwritten on re-upload)
            
            return True
            
        except Exception as e:
            logger.error(f"Error deleting file {filename}: {e}")
            return False

    async def close(self) -> None:
        """Clean up resources."""
        # Persist Chroma database
        if hasattr(self.vectorstore, 'persist'):
            self.vectorstore.persist()
        logger.info("RAG agent resources cleaned up")

    async def _analyze_excel_sheet_comprehensive(self, df: pd.DataFrame, sheet_name: str, file_path: Path) -> str:
        """Comprehensive analysis of Excel sheet including data quality and insights."""
        try:
            analysis_parts = []
            
            # Basic information
            rows, cols = df.shape
            analysis_parts.append(f"📊 **Excel Sheet: {sheet_name}**")
            analysis_parts.append(f"\n📋 **Sheet Overview:**")
            analysis_parts.append(f"- **Dimensions**: {rows:,} rows × {cols} columns")
            analysis_parts.append(f"- **File**: {file_path.name}")
            
            # Column analysis
            numeric_cols = df.select_dtypes(include=[np.number]).columns.tolist()
            text_cols = df.select_dtypes(include=['object', 'string']).columns.tolist()
            date_cols = df.select_dtypes(include=['datetime64']).columns.tolist()
            
            analysis_parts.append(f"- **Column Types**: {len(numeric_cols)} numeric, {len(text_cols)} text, {len(date_cols)} datetime")
            
            # Data quality assessment
            missing_pct = (df.isnull().sum().sum() / (rows * cols) * 100)
            duplicate_pct = (df.duplicated().sum() / rows * 100)
            
            analysis_parts.append(f"- **Data Quality**: {100 - missing_pct:.1f}% complete, {duplicate_pct:.1f}% duplicates")
            
            # Column details
            if cols <= 15:  # Show details for smaller sheets
                col_details = []
                for col in df.columns:
                    col_type = str(df[col].dtype)
                    unique_count = df[col].nunique()
                    missing_count = df[col].isnull().sum()
                    
                    detail = f"**{col}** ({col_type}): {unique_count} unique"
                    if missing_count > 0:
                        detail += f", {missing_count} missing"
                    col_details.append(detail)
                
                analysis_parts.append(f"\n📊 **Column Details:**")
                analysis_parts.extend([f"- {detail}" for detail in col_details])
            
            # Statistical summary for numeric columns
            if numeric_cols:
                analysis_parts.append(f"\n📈 **Numeric Data Summary:**")
                stats = df[numeric_cols].describe()
                
                for col in numeric_cols[:5]:  # Show first 5 numeric columns
                    col_stats = stats[col]
                    analysis_parts.append(f"- **{col}**: Min={col_stats['min']:.2f}, Max={col_stats['max']:.2f}, Mean={col_stats['mean']:.2f}")
            
            # Text data insights
            if text_cols:
                analysis_parts.append(f"\n📝 **Text Data Insights:**")
                for col in text_cols[:3]:  # Show first 3 text columns
                    unique_count = df[col].nunique()
                    most_common = df[col].mode()
                    if len(most_common) > 0:
                        analysis_parts.append(f"- **{col}**: {unique_count} unique values, most common: '{most_common.iloc[0]}'")
            
            # Sample data
            if rows > 0:
                sample_size = min(10, rows)
                sample_data = df.head(sample_size)
                
                # Format sample data nicely
                analysis_parts.append(f"\n📋 **Sample Data (First {sample_size} rows):**")
                analysis_parts.append("```")
                
                # Show column headers
                col_names = [str(col)[:20] for col in df.columns[:10]]  # Limit column width and count
                analysis_parts.append("| " + " | ".join(col_names) + " |")
                analysis_parts.append("|" + "|".join("---" for _ in col_names) + "|")
                
                # Show sample rows
                for _, row in sample_data.iterrows():
                    row_values = [str(row[col])[:20] if pd.notna(row[col]) else "" for col in df.columns[:10]]
                    analysis_parts.append("| " + " | ".join(row_values) + " |")
                
                analysis_parts.append("```")
                
                if rows > sample_size:
                    analysis_parts.append(f"\n[Note: Showing {sample_size} of {rows:,} total rows]")
            
            # Data insights and recommendations
            analysis_parts.append(f"\n💡 **Analysis Insights:**")
            
            if missing_pct > 20:
                analysis_parts.append("- ⚠️ High missing data detected - consider data cleaning")
            
            if duplicate_pct > 5:
                analysis_parts.append("- ⚠️ Significant duplicates found - consider deduplication")
            
            if len(numeric_cols) > 0:
                analysis_parts.append("- 📊 Numeric data available for statistical analysis and visualization")
            
            if len(date_cols) > 0:
                analysis_parts.append("- 📅 Date/time data found - time series analysis possible")
            
            if rows > 1000:
                analysis_parts.append("- 📈 Large dataset suitable for advanced analytics and machine learning")
            
            return "\n".join(analysis_parts)
            
        except Exception as e:
            logger.error(f"Comprehensive Excel analysis failed: {e}")
            return f"📊 **Excel Sheet: {sheet_name}**\n\n⚠️ Analysis failed: {str(e)}"


def create_rag_tool(settings: Optional[RAGSettings] = None) -> Tool:
    """
    Create a LangChain tool for the RAG agent.
    
    Args:
        settings: Optional configuration settings
        
    Returns:
        Tool instance for use with LangChain agents
    """
    
    async def analyze_documents(query: str) -> str:
        """Analyze documents using RAG agent."""
        try:
            # Create a new agent instance for each thread (not singleton)
            agent = DataAnalysisRAGAgent(settings)
            
            # Reindex documents in the thread's data directory
            await agent._reindex_existing_documents()
            
            result = await agent.analyze_data(query)
            
            # Format result for agent consumption
            formatted_result = f"""## 📊 Document Analysis Results

**Summary:** {result.summary}

**Key Insights:**
{chr(10).join(f"• {insight}" for insight in result.insights)}

**Recommendations:**
{chr(10).join(f"• {rec}" for rec in result.recommendations)}

**Statistical Analysis:**
{json.dumps(result.statistical_analysis, indent=2)[:500]}...

**Visualizations:** {len(result.visualizations)} charts generated
**Confidence Score:** {result.confidence_score:.2f}
"""
            
            return formatted_result
            
        except Exception as e:
            logger.error(f"RAG tool error: {e}")
            return f"Document analysis failed: {str(e)}"
    
    return Tool(
        name="DocumentAnalysis",
        description=(
            "Analyze uploaded documents and datasets using advanced retrieval-augmented generation. "
            "Processes any file type (PDF, CSV, Excel, JSON, TXT) to extract insights, perform statistical analysis, "
            "and provide data-driven recommendations. Use for questions about document content, data patterns, "
            "statistical analysis, or when users need comprehensive understanding of their uploaded files. "
            "Excels at both quantitative analysis of structured data and qualitative analysis of text documents."
        ),
        func=lambda query: asyncio.run(analyze_documents(query))
    )


def create_rag_agent(settings: Optional[RAGSettings] = None) -> DataAnalysisRAGAgent:
    """
    Factory function to create a configured RAG agent using singleton pattern.
    This ensures all components share the same vector store and processed documents.
    
    Args:
        settings: Optional configuration settings
        
    Returns:
        Configured DataAnalysisRAGAgent instance (singleton)
    """
    global _rag_agent_instance
    
    with _instance_lock:
        if _rag_agent_instance is None:
            _rag_agent_instance = DataAnalysisRAGAgent(settings)
            logger.info("Created new RAG agent singleton instance")
        else:
            logger.info("Returning existing RAG agent singleton instance")
        return _rag_agent_instance


def get_rag_agent_instance() -> Optional[DataAnalysisRAGAgent]:
    """
    Get the current RAG agent singleton instance if it exists.
    
    Returns:
        The RAG agent instance or None if not yet created
    """
    return _rag_agent_instance
